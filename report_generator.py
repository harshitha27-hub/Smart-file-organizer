"""
Report Generator Module
=======================
Generates Excel (.xlsx), CSV, and PDF reports covering:
  - File statistics
  - Duplicate files
  - Storage analysis
  - Prediction results
"""

import os
import csv
from datetime import datetime
from typing import Dict, Any, Optional

from modules.database import DatabaseManager
from modules.logger import AppLogger

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reports")


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


class ReportGenerator:
    """Generates Excel, CSV, and PDF reports from database data."""

    def __init__(self):
        self.db = DatabaseManager()
        self.log = AppLogger()
        os.makedirs(REPORTS_DIR, exist_ok=True)

    # ── Excel ──────────────────────────────────────────────────────────────────

    def generate_excel(self, predictions: Optional[Dict] = None) -> str:
        """Create a multi-sheet Excel workbook and return its path."""
        try:
            import openpyxl
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                         Border, Side)
            from openpyxl.utils import get_column_letter
        except ImportError:
            self.log.error("openpyxl not installed – Excel report skipped.")
            return ""

        wb = openpyxl.Workbook()

        # ── Styles ──────────────────────────────────────────────────────────
        HDR_FILL = PatternFill("solid", fgColor="1A1A2E")
        HDR_FONT = Font(bold=True, color="6C63FF", size=11)
        ALT_FILL = PatternFill("solid", fgColor="0D0D1A")
        BORDER   = Border(
            bottom=Side(style="thin", color="2A2A4A"),
        )

        def _write_sheet(ws, title: str, headers: list, rows: list):
            ws.title = title
            ws.sheet_view.showGridLines = False

            # Header row
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = HDR_FONT
                cell.fill = HDR_FILL
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            for r_idx, row in enumerate(rows, 2):
                for c_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    if r_idx % 2 == 0:
                        cell.fill = ALT_FILL
                    cell.border = BORDER

            # Auto-width
            for col in ws.columns:
                max_len = max(
                    (len(str(c.value)) for c in col if c.value), default=8
                )
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

        # ── Sheet 1: All files ──────────────────────────────────────────────
        ws1 = wb.active
        all_files = self.db.get_all_files()
        headers = ["File Name", "Category", "Size (KB)", "Date Added", "Date Modified", "Path"]
        rows = [
            (f["file_name"], f["category"],
             round(f["size"] / 1024, 1),
             f["date_added"][:19] if f["date_added"] else "",
             f["date_modified"][:19] if f["date_modified"] else "",
             f["file_path"])
            for f in all_files
        ]
        _write_sheet(ws1, "All Files", headers, rows)

        # ── Sheet 2: Category stats ─────────────────────────────────────────
        ws2 = wb.create_sheet()
        cat_rows = self.db.get_category_stats()
        _write_sheet(ws2, "Category Stats",
                     ["Category", "File Count", "Total Size (MB)"],
                     [(r[0], r[1], round((r[2] or 0) / 1024 / 1024, 2)) for r in cat_rows])

        # ── Sheet 3: Duplicates ─────────────────────────────────────────────
        ws3 = wb.create_sheet()
        dups = self.db.get_duplicates()
        _write_sheet(ws3, "Duplicates",
                     ["Original", "Duplicate", "Size (KB)", "Detected At"],
                     [(d["original_path"], d["duplicate_path"],
                       round(d["size"] / 1024, 1), d["detected_at"][:19])
                      for d in dups])

        # ── Sheet 4: Predictions ────────────────────────────────────────────
        if predictions and predictions.get("horizons"):
            ws4 = wb.create_sheet("Predictions")
            _write_sheet(ws4, "Predictions",
                         ["Horizon", "Predicted Size (MB)", "Predicted File Count"],
                         [(h, v["size_mb"], v["file_count"])
                          for h, v in predictions["horizons"].items()])

        path = os.path.join(REPORTS_DIR, f"report_{_ts()}.xlsx")
        wb.save(path)
        self.log.info(f"Excel report saved: {path}")
        return path

    # ── CSV ────────────────────────────────────────────────────────────────────

    def generate_csv(self) -> str:
        """Generate a CSV of all files and return its path."""
        path = os.path.join(REPORTS_DIR, f"files_{_ts()}.csv")
        all_files = self.db.get_all_files()
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=["file_name", "file_path", "category", "size",
                            "date_added", "date_modified", "hash_value", "is_duplicate"]
            )
            writer.writeheader()
            writer.writerows(all_files)
        self.log.info(f"CSV report saved: {path}")
        return path

    # ── PDF ────────────────────────────────────────────────────────────────────

    def generate_pdf(self, predictions: Optional[Dict] = None) -> str:
        """Generate a PDF summary report and return its path."""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import (SimpleDocTemplate, Paragraph,
                                            Spacer, Table, TableStyle)
            from reportlab.lib.units import cm
        except ImportError:
            self.log.error("reportlab not installed – PDF report skipped.")
            return ""

        path = os.path.join(REPORTS_DIR, f"report_{_ts()}.pdf")
        doc  = SimpleDocTemplate(path, pagesize=A4,
                                 leftMargin=2*cm, rightMargin=2*cm,
                                 topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        story  = []

        DARK   = colors.HexColor("#1A1A2E")
        ACCENT = colors.HexColor("#6C63FF")
        LIGHT  = colors.HexColor("#E0E0E0")

        title_style = ParagraphStyle(
            "Title", parent=styles["Title"],
            textColor=ACCENT, fontSize=22, spaceAfter=6
        )
        h2_style = ParagraphStyle(
            "H2", parent=styles["Heading2"],
            textColor=ACCENT, fontSize=14, spaceBefore=12, spaceAfter=4
        )
        body_style = ParagraphStyle(
            "Body", parent=styles["Normal"],
            fontSize=10, textColor=colors.black
        )

        def _table(headers, data, col_widths=None):
            tbl = Table([headers] + data, colWidths=col_widths)
            tbl.setStyle(TableStyle([
                ("BACKGROUND",  (0, 0), (-1, 0), DARK),
                ("TEXTCOLOR",   (0, 0), (-1, 0), ACCENT),
                ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",    (0, 0), (-1, 0), 10),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F0FA")]),
                ("GRID",        (0, 0), (-1, -1), 0.3, colors.lightgrey),
                ("FONTSIZE",    (0, 1), (-1, -1), 9),
                ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
                ("ROWPADDING",  (0, 0), (-1, -1), 4),
            ]))
            return tbl

        # Title
        story.append(Paragraph("Smart File Organizer – Report", title_style))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", body_style))
        story.append(Spacer(1, 0.4*cm))

        # Summary
        stats = self.db.get_total_stats()
        story.append(Paragraph("Summary Statistics", h2_style))
        summary_data = [
            ["Total Files",    str(stats.get("total_files", 0))],
            ["Total Size",     f"{round(stats.get('total_size', 0) / 1024 / 1024, 2)} MB"],
            ["Categories",     str(stats.get("categories", 0))],
            ["Duplicate Files",str(stats.get("duplicates", 0))],
        ]
        story.append(_table(["Metric", "Value"], summary_data, [8*cm, 8*cm]))
        story.append(Spacer(1, 0.4*cm))

        # Category breakdown
        story.append(Paragraph("Category Breakdown", h2_style))
        cat_rows = self.db.get_category_stats()
        if cat_rows:
            story.append(_table(
                ["Category", "Files", "Size (MB)"],
                [(r[0], r[1], round((r[2] or 0) / 1024 / 1024, 2)) for r in cat_rows],
                [6*cm, 4*cm, 6*cm]
            ))
        story.append(Spacer(1, 0.4*cm))

        # Duplicates
        story.append(Paragraph("Duplicate Files", h2_style))
        dups = self.db.get_duplicates()[:20]  # Cap at 20 for readability
        if dups:
            story.append(_table(
                ["Original", "Duplicate", "Size (KB)"],
                [(os.path.basename(d["original_path"]),
                  os.path.basename(d["duplicate_path"]),
                  round(d["size"] / 1024, 1)) for d in dups],
                [6*cm, 6*cm, 4*cm]
            ))
        else:
            story.append(Paragraph("No duplicates found.", body_style))
        story.append(Spacer(1, 0.4*cm))

        # Predictions
        if predictions and predictions.get("horizons"):
            story.append(Paragraph("Storage Predictions", h2_style))
            pred_data = [
                (h, f"{v['size_mb']} MB", str(v['file_count']))
                for h, v in predictions["horizons"].items()
            ]
            story.append(_table(
                ["Horizon", "Predicted Size", "Predicted Files"],
                pred_data, [6*cm, 5*cm, 5*cm]
            ))

        doc.build(story)
        self.log.info(f"PDF report saved: {path}")
        return path
